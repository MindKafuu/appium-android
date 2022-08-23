from appium.webdriver.common.appiumby import AppiumBy
from openpyxl import load_workbook

from pages.base_page import Page

class LoginPage(Page):
    USERNAME_FIELD = (AppiumBy.XPATH, '//android.widget.EditText[@index="0"]')
    PASSWORD_FEILD = (AppiumBy.XPATH, '//android.widget.EditText[@index="1"]')
    LOGIN_RESULT = (AppiumBy.XPATH, "//android.view.View[@content-desc='กรอกข้อมูลถูก เข้าสู้ระบบนะจ๊ะ']")
    FILE_PATH = "features/resources/mock_account_10.xlsx"
    USERNAME = ""
    PASSWORD = ""
    
    def input_username(self, username: str):
        self.USERNAME = username
        self.click(*self.USERNAME_FIELD)
        self.input(username, *self.USERNAME_FIELD)

    def input_password(self, password: str):
        self.PASSWORD = password
        self.click(*self.PASSWORD_FEILD)
        self.input(password, *self.PASSWORD_FEILD)

    def verify_login(self, round: int):
        result = self.find_elements(*self.LOGIN_RESULT)

        file = load_workbook(self.FILE_PATH)
        sheet_names = file.sheetnames
        current_sheet = file[sheet_names[0]]

        if len(result) > 0:
            for row in range(2, current_sheet.max_row + 1):
                if str(current_sheet[row][0].value) == self.USERNAME and str(current_sheet[row][1].value) == self.PASSWORD:
                    current_sheet[row][1 + round].value = "passed"
            
            file.save(self.FILE_PATH)
        else:
            for row in range(2, current_sheet.max_row + 1):
                if str(current_sheet[row][0].value) == self.USERNAME and str(current_sheet[row][1].value) == self.PASSWORD:
                    current_sheet[row][1 + round].value = "failed"
            
            file.save(self.FILE_PATH)

        if round == 3:
            assert len(result) > 0, "Incorrect account or password"

